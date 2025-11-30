import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import sys

def get_resource_path(relative_path):
    """获取资源的绝对路径。用于PyInstaller打包后找到资源文件"""
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller创建的临时文件夹
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel文件比较工具 - 多工作表支持")
        self.root.geometry("800x600")
        
        # 文件路径变量
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        
        self.setup_ui()
    
    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 文件1选择
        ttk.Label(file_frame, text="文件1:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        ttk.Entry(file_frame, textvariable=self.file1_path, width=60).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(file_frame, text="浏览", command=self.browse_file1).grid(row=0, column=2)
        
        # 文件2选择
        ttk.Label(file_frame, text="文件2:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(10, 0))
        ttk.Entry(file_frame, textvariable=self.file2_path, width=60).grid(row=1, column=1, padx=(0, 5), pady=(10, 0))
        ttk.Button(file_frame, text="浏览", command=self.browse_file2).grid(row=1, column=2, pady=(10, 0))
        
        # 比较按钮
        ttk.Button(main_frame, text="开始比较", command=self.compare_files).pack(pady=10)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(main_frame, text="比较结果", padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文本框和滚动条
        self.text_result = tk.Text(result_frame, wrap=tk.WORD, width=80, height=20)
        scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.text_result.yview)
        self.text_result.configure(yscrollcommand=scrollbar.set)
        
        self.text_result.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
    
    def browse_file1(self):
        filename = filedialog.askopenfilename(
            title="选择第一个Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file1_path.set(filename)
    
    def browse_file2(self):
        filename = filedialog.askopenfilename(
            title="选择第二个Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file2_path.set(filename)
    
    def log(self, message):
        """在文本框中添加日志信息"""
        self.text_result.insert(tk.END, message + "\n")
        self.text_result.see(tk.END)
        self.root.update()
    
    def clear_log(self):
        """清空日志"""
        self.text_result.delete(1.0, tk.END)
    
    def get_all_sheets(self, file_path):
        """获取Excel文件中的所有工作表名称"""
        try:
            if file_path.endswith('.xlsx'):
                wb = load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names
            else:
                # 对于.xls文件，使用pandas获取工作表名称
                xl = pd.ExcelFile(file_path)
                return xl.sheet_names
        except Exception as e:
            self.log(f"读取工作表名称时出错: {str(e)}")
            return []
    
    def compare_files(self):
        """比较两个Excel文件的所有工作表"""
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()
        
        if not file1 or not file2:
            messagebox.showerror("错误", "请选择两个Excel文件")
            return
        
        if not os.path.exists(file1) or not os.path.exists(file2):
            messagebox.showerror("错误", "选择的文件不存在")
            return
        
        try:
            self.clear_log()
            self.status_var.set("正在读取文件...")
            self.log("开始比较Excel文件...")
            self.log(f"文件1: {file1}")
            self.log(f"文件2: {file2}")
            
            # 获取所有工作表名称
            sheets1 = self.get_all_sheets(file1)
            sheets2 = self.get_all_sheets(file2)
            
            self.log(f"文件1的工作表: {', '.join(sheets1)}")
            self.log(f"文件2的工作表: {', '.join(sheets2)}")
            
            # 找出共同的工作表和独有的工作表
            common_sheets = set(sheets1) & set(sheets2)
            only_in_file1 = set(sheets1) - set(sheets2)
            only_in_file2 = set(sheets2) - set(sheets1)
            
            if only_in_file1:
                self.log(f"仅在文件1中存在的工作表: {', '.join(only_in_file1)}")
            if only_in_file2:
                self.log(f"仅在文件2中存在的工作表: {', '.join(only_in_file2)}")
            
            # 用于存储所有工作表的比较结果
            all_results = []
            
            # 比较共同的工作表
            for sheet_name in common_sheets:
                self.log(f"\n比较工作表: {sheet_name}")
                self.status_var.set(f"正在比较工作表: {sheet_name}")
                
                # 读取工作表
                df1 = pd.read_excel(file1, sheet_name=sheet_name)
                df2 = pd.read_excel(file2, sheet_name=sheet_name)
                
                self.log(f"  - 文件1 '{sheet_name}' 行数: {len(df1)}, 列数: {len(df1.columns)}")
                self.log(f"  - 文件2 '{sheet_name}' 行数: {len(df2)}, 列数: {len(df2.columns)}")
                
                # 找出差异
                result_df = self.find_differences(df1, df2, sheet_name)
                
                if len(result_df) > 0:
                    all_results.append(result_df)
                    self.log(f"  - 发现差异: {len(result_df)} 行")
                else:
                    self.log(f"  - 无差异")
            
            # 处理仅在文件1中存在的工作表
            for sheet_name in only_in_file1:
                self.log(f"\n处理仅在文件1中存在的工作表: {sheet_name}")
                df1 = pd.read_excel(file1, sheet_name=sheet_name)
                
                # 标记所有行为"仅存在于文件1"
                result_df = self.mark_all_rows_as_different(df1, sheet_name, "仅存在于文件1")
                all_results.append(result_df)
                self.log(f"  - 标记所有 {len(result_df)} 行为仅存在于文件1")
            
            # 处理仅在文件2中存在的工作表
            for sheet_name in only_in_file2:
                self.log(f"\n处理仅在文件2中存在的工作表: {sheet_name}")
                df2 = pd.read_excel(file2, sheet_name=sheet_name)
                
                # 标记所有行为"仅存在于文件2"
                result_df = self.mark_all_rows_as_different(df2, sheet_name, "仅存在于文件2")
                all_results.append(result_df)
                self.log(f"  - 标记所有 {len(result_df)} 行为仅存在于文件2")
            
            self.status_var.set("正在生成报告...")
            
            # 保存结果
            if all_results:
                # 合并所有结果
                final_result = pd.concat(all_results, ignore_index=True)
                output_file = self.save_result(final_result, file2)
                
                self.log(f"\n比较完成！")
                self.log(f"差异报告已保存至: {output_file}")
                self.log(f"总差异行数: {len(final_result)}")
                
                self.status_var.set("比较完成")
                messagebox.showinfo("完成", f"比较完成！\n差异报告已保存至: {output_file}")
            else:
                self.log(f"\n比较完成！")
                self.log("两个文件内容完全一致，无差异")
                self.status_var.set("比较完成 - 无差异")
                messagebox.showinfo("完成", "两个文件内容完全一致，无差异")
            
        except Exception as e:
            error_msg = f"比较过程中发生错误: {str(e)}"
            self.log(error_msg)
            messagebox.showerror("错误", error_msg)
            self.status_var.set("错误")
    
    def mark_all_rows_as_different(self, df, sheet_name, diff_type):
        """将整个DataFrame标记为差异"""
        data_columns = df.columns.tolist()
        
        # 创建结果行
        result_rows = []
        for idx, (_, row) in enumerate(df.iterrows()):
            result_row = {
                '工作表': sheet_name,
                '文件1行号': row.get('_original_row', idx + 1) if diff_type == "仅存在于文件1" else '',
                '文件2行号': row.get('_original_row', idx + 1) if diff_type == "仅存在于文件2" else '',
                '差异类型': diff_type
            }
            
            # 添加数据列
            for col in data_columns:
                result_row[col] = row[col]
            
            result_rows.append(result_row)
        
        # 创建结果DataFrame
        result_columns = ['工作表', '文件1行号', '文件2行号', '差异类型'] + data_columns
        return pd.DataFrame(result_rows, columns=result_columns)
    
    def find_differences(self, df1, df2, sheet_name):
        """找出两个DataFrame之间的差异"""
        # 获取列名
        data_columns = df1.columns.tolist()
        
        # 添加原始行号
        df1_with_index = df1.copy()
        df2_with_index = df2.copy()
        df1_with_index['_original_row'] = range(1, len(df1) + 1)
        df2_with_index['_original_row'] = range(1, len(df2) + 1)
        
        # 初始化匹配标记列
        df1_with_index['_matched'] = False
        df2_with_index['_matched'] = False
        
        # 找出完全相同的行
        identical_rows = []
        for idx1, row1 in df1_with_index.iterrows():
            for idx2, row2 in df2_with_index.iterrows():
                # 如果row2已经被匹配过，跳过
                if df2_with_index.at[idx2, '_matched']:
                    continue
                    
                # 检查所有单元格是否完全相同
                all_match = True
                for col in data_columns:
                    if str(row1[col]) != str(row2[col]):
                        all_match = False
                        break
                
                if all_match:
                    identical_rows.append((idx1, idx2))
                    # 标记为已匹配
                    df1_with_index.at[idx1, '_matched'] = True
                    df2_with_index.at[idx2, '_matched'] = True
                    break
        
        # 找出未匹配的行（可能的缺失行）
        df1_unmatched = df1_with_index[df1_with_index['_matched'] == False].copy()
        df2_unmatched = df2_with_index[df2_with_index['_matched'] == False].copy()
        
        # 在可能的缺失行中查找相似行（超过半数的单元格内容相同）
        similar_rows = []
        matched_in_df1 = set()
        matched_in_df2 = set()
        
        for idx1, row1 in df1_unmatched.iterrows():
            # 如果这行已经匹配，跳过
            if idx1 in matched_in_df1:
                continue
                
            best_match_idx = None
            best_match_score = 0
            threshold = len(data_columns) // 2 + 1  # 超过半数
            
            for idx2, row2 in df2_unmatched.iterrows():
                # 如果这行已经匹配，跳过
                if idx2 in matched_in_df2:
                    continue
                    
                # 计算相同单元格的数量
                match_count = 0
                for col in data_columns:
                    if str(row1[col]) == str(row2[col]):
                        match_count += 1
                
                # 如果超过阈值且比之前的匹配更好，则更新最佳匹配
                if match_count >= threshold and match_count > best_match_score:
                    best_match_score = match_count
                    best_match_idx = idx2
            
            # 如果找到相似的行，则比较差异
            if best_match_idx is not None:
                row2 = df2_unmatched.loc[best_match_idx]
                
                # 创建合并行，显示差异
                merged_row = {
                    '工作表': sheet_name,
                    '文件1行号': row1['_original_row'],
                    '文件2行号': row2['_original_row'],
                    '差异类型': '内容不同'
                }
                
                # 添加数据列，对于有差异的列显示变化
                for col in data_columns:
                    if str(row1[col]) != str(row2[col]):
                        merged_row[col] = f"{row1[col]} → {row2[col]}"
                    else:
                        merged_row[col] = row1[col]
                
                similar_rows.append(merged_row)
                
                # 记录已经匹配的行
                matched_in_df1.add(idx1)
                matched_in_df2.add(best_match_idx)
        
        # 找出真正缺失的行（在相似行匹配后仍然未匹配的行）
        only_in_file1 = []
        for idx, row in df1_unmatched.iterrows():
            if idx not in matched_in_df1:
                only_in_file1.append(row)
        
        only_in_file2 = []
        for idx, row in df2_unmatched.iterrows():
            if idx not in matched_in_df2:
                only_in_file2.append(row)
        
        # 创建结果DataFrame
        result_columns = ['工作表', '文件1行号', '文件2行号', '差异类型'] + data_columns
        
        # 处理只存在于文件1的行
        result_rows = []
        for row in only_in_file1:
            result_row = {
                '工作表': sheet_name,
                '文件1行号': row['_original_row'],
                '文件2行号': '',
                '差异类型': '仅存在于文件1'
            }
            for col in data_columns:
                result_row[col] = row[col]
            result_rows.append(result_row)
        
        # 处理只存在于文件2的行
        for row in only_in_file2:
            result_row = {
                '工作表': sheet_name,
                '文件1行号': '',
                '文件2行号': row['_original_row'],
                '差异类型': '仅存在于文件2'
            }
            for col in data_columns:
                result_row[col] = row[col]
            result_rows.append(result_row)
        
        # 添加内容不同的行
        result_rows.extend(similar_rows)
        
        # 创建最终结果DataFrame
        result_df = pd.DataFrame(result_rows, columns=result_columns)
        
        self.log(f"  - 完全相同行数: {len(identical_rows)}")
        self.log(f"  - 仅存在于文件1行数: {len(only_in_file1)}")
        self.log(f"  - 仅存在于文件2行数: {len(only_in_file2)}")
        self.log(f"  - 内容不同行数: {len(similar_rows)}")
        
        return result_df
    
    def save_result(self, result_df, file2_path):
        """保存比较结果到Excel文件"""
        # 获取文件2所在目录
        file2_dir = os.path.dirname(file2_path)
        file2_name = os.path.splitext(os.path.basename(file2_path))[0]
        
        # 生成输出文件名
        output_file = os.path.join(file2_dir, f"{file2_name}_比较结果.xlsx")
        
        # 如果文件已存在，添加序号
        counter = 1
        base_output_file = output_file
        while os.path.exists(output_file):
            output_file = os.path.join(file2_dir, f"{file2_name}_比较结果({counter}).xlsx")
            counter += 1
        
        # 使用openpyxl创建Workbook以便添加样式
        wb = Workbook()
        ws = wb.active
        ws.title = "比较结果"
        
        # 获取所有列名
        all_columns = result_df.columns.tolist()
        data_columns = [col for col in all_columns if col not in ['工作表', '文件1行号', '文件2行号', '差异类型']]
        
        # 添加标题行
        for col_idx, header in enumerate(all_columns, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 定义填充样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 内容不同
        light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 浅绿色 - 仅存在于文件1
        light_blue_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")  # 浅蓝色 - 仅存在于文件2
        
        # 添加数据行
        for row_idx, (_, row) in enumerate(result_df.iterrows(), 2):
            for col_idx, col_name in enumerate(all_columns, 1):
                cell_value = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # 根据差异类型应用不同的填充样式
                if row['差异类型'] == '内容不同':
                    # 只对内容不同的单元格标记黄色
                    if col_name in data_columns and '→' in str(cell_value):
                        cell.fill = yellow_fill
                elif row['差异类型'] == '仅存在于文件1':
                    # 对仅存在于文件1的行，用浅绿色填充文件1行号
                    if col_name == '文件1行号':
                        cell.fill = light_green_fill
                elif row['差异类型'] == '仅存在于文件2':
                    # 对仅存在于文件2的行，用浅蓝色填充文件2行号
                    if col_name == '文件2行号':
                        cell.fill = light_blue_fill
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # 限制最大宽度
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        return output_file

def main():
    root = tk.Tk()
    app = ExcelComparator(root)
    root.mainloop()

if __name__ == "__main__":
    main()