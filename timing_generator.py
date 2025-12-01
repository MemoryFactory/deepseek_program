import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import matplotlib
import sys

# 设置matplotlib使用中文字体
if sys.platform.startswith('win'):  # Windows系统
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
elif sys.platform.startswith('darwin'):  # macOS系统
    plt.rcParams['font.sans-serif'] = ['Heiti TC', 'Arial Unicode MS', 'STHeiti']
else:  # Linux系统
    plt.rcParams['font.sans-serif'] = ['WenQuanYi Micro Hei', 'AR PL UMing CN', 'Arial Unicode MS']

plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

class TimingDiagramGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("时序图生成器")
        self.root.geometry("1000x700")
        
        self.file_path = None
        self.df = None
        self.signals = []
        
        self.setup_ui()
        
    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件操作", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(file_frame, text="选择Excel文件", command=self.load_file).grid(row=0, column=0, padx=5)
        self.file_label = ttk.Label(file_frame, text="未选择文件")
        self.file_label.grid(row=0, column=1, padx=5)
        
        # 参数设置区域
        param_frame = ttk.LabelFrame(main_frame, text="参数设置", padding="5")
        param_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(param_frame, text="信号间距:").grid(row=0, column=0, padx=5)
        self.signal_spacing = tk.DoubleVar(value=1.5)
        ttk.Entry(param_frame, textvariable=self.signal_spacing, width=10).grid(row=0, column=1, padx=5)
        
        ttk.Label(param_frame, text="时间单位宽度:").grid(row=0, column=2, padx=5)
        self.time_unit_width = tk.DoubleVar(value=1.0)
        ttk.Entry(param_frame, textvariable=self.time_unit_width, width=10).grid(row=0, column=3, padx=5)
        
        ttk.Label(param_frame, text="线宽:").grid(row=0, column=4, padx=5)
        self.line_width = tk.DoubleVar(value=2.0)
        ttk.Entry(param_frame, textvariable=self.line_width, width=10).grid(row=0, column=5, padx=5)
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="生成时序图", command=self.generate_diagram).grid(row=0, column=0, padx=5)
        ttk.Button(button_frame, text="保存为JPG", command=self.save_jpg).grid(row=0, column=1, padx=5)
        ttk.Button(button_frame, text="保存为Excel", command=self.save_excel).grid(row=0, column=2, padx=5)
        
        # 图形显示区域
        self.figure = Figure(figsize=(10, 6), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, main_frame)
        self.canvas.get_tk_widget().grid(row=3, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.config(text=os.path.basename(file_path))
            try:
                self.df = pd.read_excel(file_path)
                self.process_data()
                messagebox.showinfo("成功", f"成功加载文件，共{len(self.signals)}个信号")
            except Exception as e:
                messagebox.showerror("错误", f"读取文件失败: {str(e)}")
    
    def process_data(self):
        if self.df is not None:
            self.signals = []
            for index, row in self.df.iterrows():
                signal_name = str(row.iloc[0])
                # 处理各种可能的输入格式
                values = []
                for x in row.iloc[1:]:
                    if pd.isna(x):
                        values.append(0)
                    elif str(x).strip() in ['1', '高', 'H', 'High']:
                        values.append(1)
                    elif str(x).strip() in ['0', '低', 'L', 'Low']:
                        values.append(0)
                    else:
                        values.append(0)  # 默认值
                self.signals.append({
                    'name': signal_name,
                    'values': values
                })
    
    def generate_diagram(self):
        if not self.signals:
            messagebox.showwarning("警告", "请先加载Excel文件")
            return
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        signal_spacing = self.signal_spacing.get()
        time_unit_width = self.time_unit_width.get()
        line_width = self.line_width.get()
        
        # 设置坐标轴
        num_signals = len(self.signals)
        max_time_units = max(len(signal['values']) for signal in self.signals)
        
        # 设置图形范围
        ax.set_xlim(-0.5, max_time_units * time_unit_width - 0.5)
        ax.set_ylim(-0.5, num_signals * signal_spacing - 0.5)
        
        # 绘制每个信号的时序图
        for i, signal in enumerate(self.signals):
            y_pos = (num_signals - i - 1) * signal_spacing
            values = signal['values']
            
            # 绘制信号名称（支持中文）
            ax.text(-0.8, y_pos, signal['name'], ha='right', va='center', 
                   fontsize=12, fontweight='bold')
            
            # 绘制时序线
            for t in range(len(values)):
                x_start = t * time_unit_width
                x_end = (t + 1) * time_unit_width
                
                # 当前值
                current_value = values[t]
                
                # 绘制水平线
                y_level = y_pos + current_value * 0.3
                ax.hlines(y=y_level, xmin=x_start, xmax=x_end, 
                         linewidth=line_width, color='blue')
                
                # 绘制垂直线（值变化时）
                if t > 0 and values[t] != values[t-1]:
                    ax.vlines(x=x_start, ymin=y_pos + values[t-1] * 0.3, 
                             ymax=y_pos + values[t] * 0.3, 
                             linewidth=line_width, color='blue', linestyle='-')
                
                # 绘制时间刻度
                if t < len(values) - 1:
                    ax.vlines(x=x_end, ymin=y_level-0.1, ymax=y_level+0.1, 
                             linewidth=1, color='black', alpha=0.5)
            
            # 绘制信号标识线
            ax.hlines(y=y_pos-0.1, xmin=-0.5, xmax=max_time_units * time_unit_width - 0.5, 
                     linewidth=1, color='gray', alpha=0.3)
        
        # 设置坐标轴标签（支持中文）
        ax.set_xlabel('时间')
        ax.set_ylabel('信号')
        
        # 隐藏y轴刻度
        ax.set_yticks([])
        
        # 设置x轴刻度
        time_ticks = list(range(0, max_time_units))
        ax.set_xticks([t * time_unit_width for t in time_ticks])
        ax.set_xticklabels([f'T{t}' for t in time_ticks])
        
        # 添加网格
        ax.grid(True, alpha=0.3)
        
        # 添加标题（支持中文）
        ax.set_title('时序图', fontsize=14, fontweight='bold')
        
        self.canvas.draw()
    
    def save_jpg(self):
        if not hasattr(self, 'figure') or len(self.figure.get_axes()) == 0:
            messagebox.showwarning("警告", "请先生成时序图")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存为JPG",
            defaultextension=".jpg",
            filetypes=[("JPEG files", "*.jpg"), ("PNG files", "*.png"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # 保存时确保中文字体正确
                self.figure.savefig(file_path, dpi=300, bbox_inches='tight', 
                                  facecolor='white', edgecolor='none')
                messagebox.showinfo("成功", f"时序图已保存为: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    def save_excel(self):
        if not self.signals:
            messagebox.showwarning("警告", "请先加载Excel文件")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存为Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "时序图"
                
                # 定义边框样式
                thin_border = Side(style='thin', color='000000')
                thick_border = Side(style='thick', color='000000')
                
                # 设置标题行
                ws.cell(row=1, column=1, value="信号/时间")
                title_cell = ws.cell(row=1, column=1)
                title_cell.font = Font(bold=True)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 写入时间标签
                max_time_units = max(len(signal['values']) for signal in self.signals)
                for t in range(max_time_units):
                    time_cell = ws.cell(row=1, column=t+2, value=f'T{t}')
                    time_cell.font = Font(bold=True)
                    time_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 写入信号数据，每个信号间隔一行
                current_row = 2  # 从第2行开始
                
                for i, signal in enumerate(self.signals):
                    values = signal['values']
                    
                    # 写入信号名称
                    name_cell = ws.cell(row=current_row, column=1, value=signal['name'])
                    name_cell.font = Font(bold=True)
                    name_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 写入时序数据并设置边框
                    for t in range(len(values)):
                        cell = ws.cell(row=current_row, column=t+2, value=values[t])
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 设置边框
                        borders = {}
                        
                        # 上边框（值为1时用粗线，值为0时用细线）
                        if values[t] == 1:
                            borders['top'] = thick_border
                        else:
                            borders['top'] = thin_border
                        
                        # 下边框（值为0时用粗线，值为1时用细线）
                        if values[t] == 0:
                            borders['bottom'] = thick_border
                        else:
                            borders['bottom'] = thin_border
                        
                        # 左边框（值变化时用粗线）
                        if t > 0 and values[t] != values[t-1]:
                            borders['left'] = thick_border
                        else:
                            borders['left'] = thin_border
                        
                        # 右边框（值变化时用粗线）
                        if t < len(values)-1 and values[t] != values[t+1]:
                            borders['right'] = thick_border
                        else:
                            borders['right'] = thin_border
                        
                        cell.border = Border(**borders)
                    
                    # 移动到下一行，间隔一行
                    current_row += 2
                
                # 设置列宽
                ws.column_dimensions['A'].width = 15  # 信号名称列
                for col in range(2, max_time_units + 2):
                    col_letter = chr(64 + col)  # 转换为列字母
                    ws.column_dimensions[col_letter].width = 8
                
                # 设置行高
                for row in range(1, current_row):
                    if row % 2 == 1:  # 数据行
                        ws.row_dimensions[row].height = 25
                    else:  # 间隔行
                        ws.row_dimensions[row].height = 5
                
                # 合并间隔行的第一列单元格（为了美观）
                for row in range(3, current_row, 2):
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_time_units+1)
                    merged_cell = ws.cell(row=row, column=1)
                    merged_cell.value = ""  # 清空内容
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"时序图Excel已保存为: {file_path}")
                
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")

def main():
    root = tk.Tk()
    app = TimingDiagramGenerator(root)
    root.mainloop()

if __name__ == "__main__":
    main()